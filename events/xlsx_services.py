import openpyxl
import zipfile

from io import BytesIO 

from datetime import datetime

from django.http import HttpResponse
from django.contrib.gis.geos import Point
from django.db import transaction
from django.utils.timezone import make_aware

from venues.models import Venue
from .models import Event, EventStatus

def parse_coordinates(coord_str):
    """
    Парсит строку вида 'longitude, latitude' (например '37.61, 55.75') в Point.
    """
    try:
        lon, lat = map(float, coord_str.replace(";", ",").split(","))
        return Point(lon, lat, srid=4326)
    except (ValueError, AttributeError):
        return None

def parse_excel_date(value):
    """
    Превращает значение из Excel (datetime или строку) в Aware Datetime (UTC).
    """
    if not value:
        return None
        
    dt = value
    
    if isinstance(value, str):
        try:
            dt = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            try:
                dt = datetime.strptime(value, "%Y-%m-%d")
            except ValueError:
                return None

    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            return make_aware(dt)
        return dt
        
    return None

def export_events_to_xlsx(queryset):
    """
    Генерирует XLSX-файл из QuerySet событий.
    Возвращает HttpResponse с файлом.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"

    headers = [
        "Дата публикации", 
        "Дата начала", 
        "Дата завершения", 
        "Место проведения", 
        "Рейтинг"
    ]
    ws.append(headers)

    for event in queryset:
        ws.append([
            event.publish_at.strftime("%Y-%m-%d %H:%M") if event.publish_at else "",
            event.start_at.strftime("%Y-%m-%d %H:%M"),
            event.end_at.strftime("%Y-%m-%d %H:%M"),
            event.venue.name,
            event.rating
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()

    response = HttpResponse(
        content=content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="events_export.xlsx"'
    return response


def import_events_from_xlsx(file_obj, user):
    """
    Читает XLSX файл и создает события.
    Возвращает статистику (создано, ошибок).
    """
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except (zipfile.BadZipFile, OSError):
        return {"created": 0, "errors": ["Файл поврежден или не является корректным XLSX."]}
    ws = wb.active
    
    created_count = 0
    errors = []

    rows = ws.iter_rows(min_row=2, values_only=True)
    
    for i, row in enumerate(rows, start=2):
        try:
            with transaction.atomic():
                if not row or not row[0]:
                    continue

                title = row[0]
                description = row[1] or ""
                publish_at = row[2] 
                start_at = row[3]
                end_at = row[4]
                venue_name = row[5]
                coords_str = str(row[6])
                rating = row[7] or 0
                
                start_at = parse_excel_date(start_at)
                end_at = parse_excel_date(end_at)
                publish_at = parse_excel_date(publish_at)

                point = parse_coordinates(coords_str)
                if not point:
                    venue = Venue.objects.filter(name=venue_name).first()
                    if not venue:
                        raise ValueError(f"Venue '{venue_name}' not found and no coords")
                else:
                    venue, _ = Venue.objects.get_or_create(
                        name=venue_name,
                        defaults={"location": point}
                    )

                Event.objects.create(
                    title=title,
                    description=description,
                    publish_at=publish_at,
                    start_at=start_at,
                    end_at=end_at,
                    venue=venue,
                    rating=rating,
                    author=user,
                    status=EventStatus.DRAFT
                )
                created_count += 1

        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    return {"created": created_count, "errors": errors}
